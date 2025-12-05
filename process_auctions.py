#!/usr/bin/env python3
import asyncio
import aiohttp
import time
import re
import sqlite3
import openpyxl
from openpyxl import Workbook
from urllib.parse import quote_plus
import folium
from folium import plugins
import json
import logging
from math import radians, cos, sin, asin, sqrt

# Configure logging
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# -------------------------------------------
# CONFIG
# -------------------------------------------
GEOCODE_CACHE_DB = "geocode_cache.db"
CONCURRENT_WORKERS = 5
NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"
PHILA_AIS_URL = "https://api.phila.gov/ais_doc/v1/search"
PHILA_GATEKEEPER_KEY = "6ba4de64d6ca99aa4db3b9194e37adbf"
USER_AGENT = "AuctionProcessor/1.0 (your@email.com)"


# -------------------------------------------
# DATABASE CACHE
# -------------------------------------------
def init_cache():
    conn = sqlite3.connect(GEOCODE_CACHE_DB)
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS cache (
            query TEXT PRIMARY KEY,
            lat REAL,
            lng REAL
        )
    """)
    conn.commit()
    return conn


def cache_get(conn, query):
    cur = conn.cursor()
    cur.execute("SELECT lat, lng FROM cache WHERE query = ?", (query,))
    row = cur.fetchone()
    return row if row else None


def cache_set(conn, query, lat, lng):
    cur = conn.cursor()
    cur.execute(
        "INSERT OR REPLACE INTO cache (query, lat, lng) VALUES (?, ?, ?)",
        (query, lat, lng),
    )
    conn.commit()


def init_neighborhood_cache():
    """Initialize neighborhood cache table."""
    conn = sqlite3.connect(GEOCODE_CACHE_DB)
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS neighborhood_cache (
            lat REAL,
            lng REAL,
            neighborhood TEXT,
            PRIMARY KEY (lat, lng)
        )
    """)
    conn.commit()
    return conn


def cache_get_neighborhood(conn, lat, lng):
    """Get cached neighborhood for coordinates."""
    cur = conn.cursor()
    cur.execute(
        "SELECT neighborhood FROM neighborhood_cache WHERE lat = ? AND lng = ?",
        (lat, lng)
    )
    row = cur.fetchone()
    return row[0] if row else None


def cache_set_neighborhood(conn, lat, lng, neighborhood):
    """Cache neighborhood for coordinates."""
    cur = conn.cursor()
    cur.execute(
        "INSERT OR REPLACE INTO neighborhood_cache (lat, lng, neighborhood) VALUES (?, ?, ?)",
        (lat, lng, neighborhood),
    )
    conn.commit()


# -------------------------------------------
# NEIGHBORHOOD & CLUSTERING
# -------------------------------------------
def haversine_distance(lat1, lng1, lat2, lng2):
    """Calculate distance between two coordinates in feet."""
    # Convert decimal degrees to radians
    lat1, lng1, lat2, lng2 = map(radians, [lat1, lng1, lat2, lng2])

    # Haversine formula
    dlat = lat2 - lat1
    dlng = lng2 - lng1
    a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlng/2)**2
    c = 2 * asin(sqrt(a))
    r = 3959  # Radius of earth in miles

    return c * r * 5280  # Convert to feet


async def get_neighborhood(session, cache_conn, lat, lng):
    """Get neighborhood from coordinates using Nominatim reverse geocoding."""
    if lat is None or lng is None:
        return "Unknown"

    # Check cache first
    cached_neighborhood = cache_get_neighborhood(cache_conn, lat, lng)
    if cached_neighborhood:
        logger.debug(f"[CACHE HIT] Neighborhood for ({lat:.4f}, {lng:.4f}) -> {cached_neighborhood}")
        return cached_neighborhood

    try:
        # Use Nominatim reverse geocoding
        url = "https://nominatim.openstreetmap.org/reverse"
        params = {
            "format": "jsonv2",
            "lat": lat,
            "lon": lng,
        }

        logger.debug(f"[NOMINATIM REVERSE] Requesting {url} with params: {params}")

        async with session.get(url, params=params, timeout=5) as resp:
            logger.debug(f"[NOMINATIM REVERSE] Response status: {resp.status}")

            if resp.status == 200:
                data = await resp.json()
                logger.debug(f"[NOMINATIM REVERSE] Full response: {data}")

                address = data.get("address", {})
                logger.debug(f"[NOMINATIM REVERSE] Address object: {address}")
                logger.debug(f"[NOMINATIM REVERSE] Residential field: {address.get('residential')}")
                logger.debug(f"[NOMINATIM REVERSE] Neighbourhood field: {address.get('neighbourhood')}")

                # Prefer residential field, fall back to neighbourhood
                neighborhood = address.get("residential") or address.get("neighbourhood") or "Unknown"
                logger.info(f"[NOMINATIM REVERSE] Found neighborhood: {neighborhood} for ({lat:.4f}, {lng:.4f})")

                # Cache the result
                cache_set_neighborhood(cache_conn, lat, lng, neighborhood)
                return neighborhood
            else:
                logger.warning(f"[NOMINATIM REVERSE] Non-200 status: {resp.status}")
    except Exception as e:
        logger.error(f"[NOMINATIM REVERSE] Error getting neighborhood: {e}", exc_info=True)

    return "Unknown"


def cluster_properties(properties, max_distance_feet=300):
    """Cluster properties that are within max_distance_feet of each other."""
    if not properties:
        return []

    clusters = []
    visited = set()

    for i, prop in enumerate(properties):
        if i in visited or prop.get("lat") is None or prop.get("lng") is None:
            continue

        cluster = [prop]
        visited.add(i)

        # Find all properties within max_distance_feet
        for j, other_prop in enumerate(properties):
            if j <= i or j in visited:
                continue
            if other_prop.get("lat") is None or other_prop.get("lng") is None:
                continue

            distance = haversine_distance(
                prop["lat"], prop["lng"],
                other_prop["lat"], other_prop["lng"]
            )

            if distance <= max_distance_feet:
                cluster.append(other_prop)
                visited.add(j)

        clusters.append(cluster)

    return clusters


# -------------------------------------------
# GEOCODING
# -------------------------------------------
async def geocode_address(session, cache_conn, address, opa=None):
    """Returns (lat, lng) or (None, None). Tries OPA first, then full address, then zipcode."""

    cached = cache_get(cache_conn, address)
    if cached:
        lat, lng = cached
        if lat is None:
            logger.debug(f"[CACHE HIT] {address} -> (None, None) [cached failure]")
        else:
            logger.debug(f"[CACHE HIT] {address} -> ({lat:.4f}, {lng:.4f})")
        return cached

    # Try OPA first if available
    if opa:
        logger.debug(f"[OPA] Querying OPA {opa} for: {address}")
        params = {
            "gatekeeperKey": PHILA_GATEKEEPER_KEY,
        }
        url = f"{PHILA_AIS_URL}/{opa}"

        async with session.get(url, params=params) as resp:
            if resp.status == 200:
                data = await resp.json()
                if data.get("features") and len(data["features"]) > 0:
                    coords = data["features"][0].get("geometry", {}).get("coordinates")
                    if coords and len(coords) >= 2:
                        lng = float(coords[0])
                        lat = float(coords[1])
                        cache_set(cache_conn, address, lat, lng)
                        logger.debug(f"[OPA] Found OPA {opa} -> ({lat:.4f}, {lng:.4f})")
                        await asyncio.sleep(0.5)  # Small delay for API politeness
                        return (lat, lng)

        logger.debug(f"[OPA] No results for OPA {opa}, trying Nominatim...")

    # Fall back to full address with Nominatim
    logger.debug(f"[NOMINATIM] Querying: {address}")
    params = {
        "q": address,
        "format": "jsonv2",
        "limit": 1,
    }

    async with session.get(NOMINATIM_URL, params=params) as resp:
        if resp.status != 200:
            logger.debug(f"[NOMINATIM] Failed ({resp.status}): {address}, trying zipcode fallback...")
            return await geocode_zipcode_fallback(session, cache_conn, address, opa)

        data = await resp.json()
        if data:
            lat = float(data[0]["lat"])
            lng = float(data[0]["lon"])
            cache_set(cache_conn, address, lat, lng)
            logger.debug(f"[NOMINATIM] Found: {address} -> ({lat:.4f}, {lng:.4f})")
            await asyncio.sleep(1)  # respect Nominatim rate limit
            return (lat, lng)

    logger.debug(f"[NOMINATIM] No results: {address}, trying zipcode fallback...")
    return await geocode_zipcode_fallback(session, cache_conn, address, opa)


async def geocode_zipcode_fallback(session, cache_conn, address, opa=None):
    """Fallback: extract zipcode and try geocoding that."""
    # Extract zipcode (5 digits) from the address
    zipcode_match = re.search(r'\b(\d{5})\b', str(address))
    if not zipcode_match:
        logger.debug(f"[ZIPCODE FALLBACK] No zipcode found in: {address}")
        return (None, None)

    zipcode = zipcode_match.group(1)
    logger.debug(f"[ZIPCODE FALLBACK] Extracted zipcode {zipcode} from: {address}")

    params = {
        "q": zipcode,
        "format": "jsonv2",
        "limit": 1,
    }

    async with session.get(NOMINATIM_URL, params=params) as resp:
        if resp.status != 200:
            logger.debug(f"[ZIPCODE FALLBACK] Failed ({resp.status}): {zipcode}")
            return (None, None)

        data = await resp.json()
        if data:
            lat = float(data[0]["lat"])
            lng = float(data[0]["lon"])
            # Cache using zipcode
            cache_set(cache_conn, zipcode, lat, lng)
            logger.debug(f"[ZIPCODE FALLBACK] Found {zipcode} -> ({lat:.4f}, {lng:.4f})")
            await asyncio.sleep(1)  # respect Nominatim rate limit
            return (lat, lng)

    logger.debug(f"[ZIPCODE FALLBACK] No results for zipcode: {zipcode}")
    return (None, None)


async def geocode_opa_fallback(session, cache_conn, address, opa=None):
    """Fallback: use Philadelphia AIS API with OPA account number. Then try zipcode."""
    if not opa:
        logger.debug(f"[OPA FALLBACK] No OPA provided for: {address}, trying zipcode fallback...")
        return await geocode_zipcode_fallback(session, cache_conn, address, opa)

    logger.debug(f"[OPA FALLBACK] Querying OPA {opa} for: {address}")

    params = {
        "gatekeeperKey": PHILA_GATEKEEPER_KEY,
    }

    url = f"{PHILA_AIS_URL}/{opa}"

    async with session.get(url, params=params) as resp:
        if resp.status != 200:
            logger.debug(f"[OPA FALLBACK] Failed ({resp.status}): OPA {opa}, trying zipcode fallback...")
            return await geocode_zipcode_fallback(session, cache_conn, address, opa)

        data = await resp.json()
        if data.get("features") and len(data["features"]) > 0:
            coords = data["features"][0].get("geometry", {}).get("coordinates")
            if coords and len(coords) >= 2:
                lng = float(coords[0])
                lat = float(coords[1])
                cache_set(cache_conn, address, lat, lng)
                logger.debug(f"[OPA FALLBACK] Found OPA {opa} -> ({lat:.4f}, {lng:.4f})")
                await asyncio.sleep(0.5)  # Small delay for API politeness
                return (lat, lng)

    logger.debug(f"[OPA FALLBACK] No results for OPA: {opa}, trying zipcode fallback...")
    return await geocode_zipcode_fallback(session, cache_conn, address, opa)


# -------------------------------------------
# HELPER FUNCTIONS
# -------------------------------------------
def format_currency(value):
    """Format a value as currency, or return 'N/A' if not a number."""
    if isinstance(value, (int, float)):
        return f"${value:,.2f}"
    return "N/A"


# -------------------------------------------
# PARSE AMPERSAND ROWS
# -------------------------------------------
def split_ampersand_field(value):
    """Split 'A & B & C' into ['A','B','C']."""
    if not value:
        return []
    return [v.strip() for v in value.split("&")]


# -------------------------------------------
# MAP GENERATION
# -------------------------------------------
def create_interactive_map(results, map_path):
    """Create an interactive Folium map with neighborhood legend and proximity clustering."""
    logger.info("Creating interactive map...")

    # Filter valid results
    valid_results = [r for r in results if r["lat"] is not None and r["lng"] is not None]
    if not valid_results:
        logger.warning("No valid coordinates found for map generation")
        return

    # Calculate map center
    center_lat = sum(r["lat"] for r in valid_results) / len(valid_results)
    center_lng = sum(r["lng"] for r in valid_results) / len(valid_results)
    logger.info(f"Map center: ({center_lat:.4f}, {center_lng:.4f})")

    # Create map
    map_obj = folium.Map(
        location=[center_lat, center_lng],
        zoom_start=12,
        tiles="OpenStreetMap",
        prefer_canvas=True
    )

    # Group properties by neighborhood
    neighborhoods = {}
    for r in valid_results:
        neighborhood = r.get("neighborhood", "Unknown")
        if neighborhood not in neighborhoods:
            neighborhoods[neighborhood] = []
        neighborhoods[neighborhood].append(r)

    # Create clusters for nearby properties
    neighborhood_markers = {}
    for neighborhood, props in neighborhoods.items():
        clusters = cluster_properties(props, max_distance_feet=300)
        neighborhood_markers[neighborhood] = {
            "count": len(props),
            "clusters": clusters,
            "properties": props
        }

    logger.info(f"Found {len(neighborhoods)} neighborhoods")

    # Create feature groups for each neighborhood
    feature_groups = {}
    marker_ids = {}  # Store marker IDs for legend interaction
    marker_id_counter = 0

    for neighborhood, data in neighborhood_markers.items():
        fg = folium.FeatureGroup(name=neighborhood, show=True)
        feature_groups[neighborhood] = fg
        marker_ids[neighborhood] = {}

        # Add markers for each cluster
        for cluster_idx, cluster in enumerate(data["clusters"]):
            if len(cluster) == 1:
                # Single property
                r = cluster[0]
                popup_html = _create_popup_html(r)
                popup = folium.Popup(popup_html, max_width=350)
                marker_color, marker_icon = _get_marker_color_icon(r['status'])

                marker = folium.Marker(
                    location=[r['lat'], r['lng']],
                    popup=popup,
                    tooltip=f"{r['address']} - {r['status']}",
                    icon=folium.Icon(color=marker_color, icon=marker_icon, prefix='fa')
                )
                marker.add_to(fg)

                # Store marker ID for legend interaction
                marker_id = f"marker_{marker_id_counter}"
                marker_id_counter += 1
                if neighborhood not in marker_ids:
                    marker_ids[neighborhood] = {}
                marker_ids[neighborhood][r['address']] = marker_id

                # Add custom attribute to marker for CSS targeting
                marker.options['id'] = marker_id
            else:
                # Clustered properties - create a cluster marker
                cluster_lat = sum(p["lat"] for p in cluster) / len(cluster)
                cluster_lng = sum(p["lng"] for p in cluster) / len(cluster)

                # Create HTML for cluster popup
                cluster_html = f"""
                <div style="width: 400px; font-family: Arial, sans-serif; font-size: 12px;">
                    <h4 style="margin-top: 0; margin-bottom: 10px;">Properties Cluster ({len(cluster)} nearby)</h4>
                    <div style="max-height: 300px; overflow-y: auto;">
                """

                for prop in cluster:
                    cluster_html += f"""
                    <div style="margin-bottom: 8px; padding: 8px; background: #f9f9f9; border-left: 3px solid #007AFF;">
                        <strong>{prop['address']}</strong><br>
                        <small>Auction: {prop['auction_id']} | Status: {prop['status']}</small><br>
                        <a href="{prop['bid4assets_link']}" target="_blank" style="color: #007AFF; text-decoration: none; font-size: 11px;">
                            View Auction →
                        </a>
                    </div>
                    """

                cluster_html += """
                    </div>
                </div>
                """

                popup = folium.Popup(cluster_html, max_width=450)

                # Use a different icon for clusters
                marker = folium.Marker(
                    location=[cluster_lat, cluster_lng],
                    popup=popup,
                    tooltip=f"Cluster: {len(cluster)} properties within 300 feet",
                    icon=folium.Icon(color='red', icon='sitemap', prefix='fa')
                )
                marker.add_to(fg)

                # Store marker IDs for all properties in cluster
                marker_id = f"marker_{marker_id_counter}"
                marker_id_counter += 1
                for prop in cluster:
                    marker_ids[neighborhood][prop['address']] = marker_id
                marker.options['id'] = marker_id

        fg.add_to(map_obj)

    # Create custom legend with neighborhood counts
    legend_html = _create_legend_html(neighborhood_markers, marker_ids)
    from folium import Element
    map_obj.get_root().html.add_child(Element(legend_html))

    # Add layer control
    folium.LayerControl().add_to(map_obj)

    # Save map
    map_obj.save(map_path)
    logger.info(f"✔ Interactive map saved to: {map_path}")
    print(f"✔ Interactive map saved to: {map_path}")


def _create_popup_html(r):
    """Create popup HTML for a single property."""
    open_date_str = str(r["open_date"]) if r["open_date"] else "N/A"
    min_bid_str = format_currency(r['min_bid'])
    debt_amount_str = format_currency(r['debt_amount'])

    popup_html = f"""
    <div style="width: 320px; font-family: Arial, sans-serif; font-size: 12px;">
        <h4 style="margin-top: 0; margin-bottom: 10px;">{r['address']}</h4>

        <div style="margin-bottom: 10px;">
            <strong>Auction ID:</strong> {r['auction_id']}<br>
            <strong>Neighborhood:</strong> {r.get('neighborhood', 'Unknown')}<br>
            <strong>Status:</strong> {r['status']}<br>
            <strong>Start Price:</strong> {min_bid_str}<br>
            <strong>Auction Opens:</strong> {open_date_str}
        </div>

        <div style="margin-bottom: 10px;">
            <strong>Property Info:</strong><br>
            OPA: {r['opa'] or 'N/A'}<br>
            Book/Writ: {r['book_writ'] or 'N/A'}<br>
            Debt Amount: {debt_amount_str}
        </div>

        <div style="margin-bottom: 10px;">
            <strong>Links:</strong><br>
    """

    if r['bid4assets_link']:
        popup_html += f"<a href='{r['bid4assets_link']}' target='_blank'>🔨 Bid4Assets Auction</a><br>"

    if r['phila_link']:
        popup_html += f"<a href='{r['phila_link']}' target='_blank'>🏠 Philly OPA Record</a><br>"

    if r['streetview']:
        popup_html += f"<a href='{r['streetview']}' target='_blank'>🚗 Google Street View</a><br>"

    popup_html += """
        </div>
    </div>
    """
    return popup_html


def _get_marker_color_icon(status):
    """Determine marker color and icon based on status."""
    status = str(status).lower()
    if 'sold' in status:
        return 'green', 'check'
    elif 'withdrawn' in status or 'cancelled' in status:
        return 'gray', 'times'
    elif 'postponed' in status:
        return 'orange', 'clock'
    else:
        return 'blue', 'gavel'


def _create_legend_html(neighborhood_markers, marker_ids):
    """Create interactive HTML legend for neighborhoods."""
    legend_html = """
    <div id="legend" style="
        position: fixed;
        bottom: 50px; left: 50px;
        width: 320px;
        background-color: white;
        border: 2px solid grey;
        border-radius: 5px;
        z-index: 9999;
        font-size: 14px;
        padding: 12px;
        box-shadow: 2px 2px 6px rgba(0,0,0,0.3);
        font-family: Arial, sans-serif;
        max-height: 500px;
        overflow-y: auto;
    ">
        <h3 style="margin-top: 0; margin-bottom: 12px; font-size: 16px;">
            Properties by Neighborhood
        </h3>
    """

    # Sort neighborhoods by count (descending)
    sorted_neighborhoods = sorted(
        neighborhood_markers.items(),
        key=lambda x: x[1]['count'],
        reverse=True
    )

    for neighborhood, data in sorted_neighborhoods:
        count = data['count']
        properties = data['properties']

        # Create addresses list HTML (hidden by default)
        addresses_html = '<div style="display:none;" class="addresses-list">'
        for prop in sorted(properties, key=lambda x: x['address']):
            marker_id = marker_ids.get(neighborhood, {}).get(prop['address'], '')
            addresses_html += f"""
            <div class="address-item" data-marker-id="{marker_id}" style="padding: 6px 8px; margin: 4px 0; background: #f0f0f0; border-radius: 3px; font-size: 12px; cursor: pointer; transition: all 0.2s;" onmouseover="this.style.backgroundColor='#ddd'; highlightMarker('{marker_id}');" onmouseout="this.style.backgroundColor='#f0f0f0'; unhighlightMarker('{marker_id}');" onclick="panToMarker('{marker_id}');">
                <a href="{prop['bid4assets_link']}" target="_blank" style="color: #0066cc; text-decoration: none;" onclick="event.stopPropagation();">
                    {prop['address']}
                </a>
            </div>
            """
        addresses_html += '</div>'

        legend_html += f"""
        <div style="
            margin-bottom: 8px;
            padding: 10px;
            background-color: #f9f9f9;
            border-left: 4px solid #0066cc;
            cursor: pointer;
            border-radius: 3px;
            transition: background-color 0.2s;
        " class="neighborhood-item" onmouseover="this.style.backgroundColor='#e6f2ff'" onmouseout="this.style.backgroundColor='#f9f9f9'" onclick="this.querySelector('.addresses-list').style.display = this.querySelector('.addresses-list').style.display === 'none' ? 'block' : 'none'">
            <strong style="font-size: 13px;">{neighborhood}</strong>
            <span style="
                display: inline-block;
                background-color: #0066cc;
                color: white;
                border-radius: 12px;
                padding: 2px 8px;
                font-size: 12px;
                margin-left: 8px;
            ">{count}</span>
            {addresses_html}
        </div>
        """

    legend_html += """
        <div style="margin-top: 12px; padding-top: 12px; border-top: 1px solid #ddd; font-size: 11px; color: #666;">
            <strong>Legend:</strong><br>
            <span style="color: blue;">●</span> Blue = Active<br>
            <span style="color: green;">●</span> Green = Sold<br>
            <span style="color: orange;">●</span> Orange = Postponed<br>
            <span style="color: gray;">●</span> Gray = Withdrawn/Cancelled
        </div>
    </div>

    <style>
    .marker-highlighted {
        filter: drop-shadow(0 0 6px rgba(255, 255, 0, 0.8)) !important;
        transform: scale(1.4) !important;
    }
    </style>

    <script>
    // Global marker storage
    var markerMap = {};

    function highlightMarker(markerId) {
        if (!markerId || !markerMap[markerId]) return;
        var marker = markerMap[markerId];
        if (marker && marker._icon) {
            marker._icon.classList.add('marker-highlighted');
            marker.setZIndexOffset(1000);
        }
    }

    function unhighlightMarker(markerId) {
        if (!markerId || !markerMap[markerId]) return;
        var marker = markerMap[markerId];
        if (marker && marker._icon) {
            marker._icon.classList.remove('marker-highlighted');
            marker.setZIndexOffset(0);
        }
    }

    function panToMarker(markerId) {
        if (!markerId || !markerMap[markerId]) return;
        var marker = markerMap[markerId];
        window.map.setView(marker.getLatLng(), 16);
        setTimeout(function() {
            marker.openPopup();
        }, 300);
    }

    // Hook into map creation to store marker references
    window.addEventListener('load', function() {
        setTimeout(function() {
            if (window.map && window.map._layers) {
                for (var id in window.map._layers) {
                    var layer = window.map._layers[id];
                    if (layer instanceof L.Marker && layer.options && layer.options.id) {
                        markerMap[layer.options.id] = layer;
                    }
                }
            }
        }, 500);
    });

    document.addEventListener('DOMContentLoaded', function() {
        // Make legend draggable
        const legend = document.getElementById('legend');
        let pos1 = 0, pos2 = 0, pos3 = 0, pos4 = 0;

        const header = legend.querySelector('h3');
        header.style.cursor = 'grab';
        header.onmousedown = dragMouseDown;

        function dragMouseDown(e) {
            e.preventDefault();
            pos3 = e.clientX;
            pos4 = e.clientY;
            document.onmouseup = closeDragElement;
            document.onmousemove = elementDrag;
        }

        function elementDrag(e) {
            e.preventDefault();
            pos1 = pos3 - e.clientX;
            pos2 = pos4 - e.clientY;
            pos3 = e.clientX;
            pos4 = e.clientY;
            legend.style.bottom = (legend.offsetParent.clientHeight - legend.offsetTop - legend.offsetHeight - pos2) + 'px';
            legend.style.left = (legend.offsetLeft - pos1) + 'px';
        }

        function closeDragElement() {
            document.onmouseup = null;
            document.onmousemove = null;
        }
    });
    </script>
    """

    return legend_html


# -------------------------------------------
# MAIN PROCESSOR
# -------------------------------------------
async def process_file(input_path, output_path, geojson_path, map_path):
    logger.info(f"Starting processing: {input_path}")
    cache_conn = init_cache()
    logger.info("Geocode cache initialized")
    init_neighborhood_cache()
    logger.info("Neighborhood cache initialized")

    wb = openpyxl.load_workbook(input_path)
    sheet = wb.active
    logger.info(f"Loaded Excel file with {sheet.max_row} rows")

    # Row 3 = headers
    headers = [cell.value for cell in sheet[3]]
    logger.info(f"Found headers: {headers}")

    # Required fields
    idx_auction_id = headers.index("Auction ID")
    idx_status = headers.index("Status")
    idx_min_bid = headers.index("Minimum Bid")
    idx_open = headers.index("Bidding Open Date/Time")
    idx_attorney = headers.index("Attorney")
    idx_book = headers.index("Book/Writ")
    idx_opa = headers.index("OPA")
    idx_address = headers.index("Address")

    # Optional field - default to None if not found
    try:
        idx_debt = headers.index("Debt Amount")
        logger.info("Found 'Debt Amount' header")
    except ValueError:
        idx_debt = None
        logger.warning("'Debt Amount' header not found, defaulting all debts to $0")

    tasks = []
    results = []
    logger.info(f"Starting geocoding with {CONCURRENT_WORKERS} concurrent workers...")

    async with aiohttp.ClientSession(headers={"User-Agent": USER_AGENT}) as session:
        sem = asyncio.Semaphore(CONCURRENT_WORKERS)

        for row_num, row in enumerate(sheet.iter_rows(min_row=4), start=4):
            auction_id = str(row[idx_auction_id].value).strip()
            status = row[idx_status].value
            min_bid = row[idx_min_bid].value
            open_date = row[idx_open].value
            attorney = row[idx_attorney].value
            debt_amount = row[idx_debt].value if idx_debt is not None else 0
            book = row[idx_book].value
            opa_raw = row[idx_opa].value
            addr_raw = row[idx_address].value

            addresses = split_ampersand_field(addr_raw)
            opas = split_ampersand_field(str(opa_raw))
            books = split_ampersand_field(str(book))

            # normalize lengths
            max_len = max(len(addresses), len(opas), len(books))
            addresses += [None] * (max_len - len(addresses))
            opas += [None] * (max_len - len(opas))
            books += [None] * (max_len - len(books))

            for i in range(max_len):
                addr = addresses[i]
                opa = opas[i]
                bk = books[i]

                if not addr:
                    continue

                async def worker(addr=addr, opa=opa, bk=bk, auction_id=auction_id, status=status,
                               min_bid=min_bid, open_date=open_date, attorney=attorney, debt_amount=debt_amount):
                    async with sem:
                        lat, lng = await geocode_address(session, cache_conn, addr, opa)
                        neighborhood = "Unknown"
                        if lat and lng:
                            neighborhood = await get_neighborhood(session, cache_conn, lat, lng)
                        return {
                            "auction_id": auction_id,
                            "status": status,
                            "min_bid": min_bid,
                            "open_date": str(open_date) if open_date else None,
                            "attorney": attorney,
                            "debt_amount": debt_amount,
                            "book_writ": bk,
                            "opa": opa,
                            "address": addr,
                            "lat": lat,
                            "lng": lng,
                            "neighborhood": neighborhood,
                            "phila_link": f"https://property.phila.gov/?p={opa}" if opa else None,
                            "bid4assets_link": f"https://www.bid4assets.com/auction/index/{auction_id}",
                            "streetview": f"https://www.google.com/maps?q={addr}&layer=c" if addr else None,
                        }

                tasks.append(worker())

        logger.info(f"Created {len(tasks)} geocoding tasks, waiting for completion...")
        results = await asyncio.gather(*tasks)
        logger.info(f"Geocoding complete. Processing {len(results)} results...")

    # -------------------------------------------
    # WRITE NEW EXCEL
    # -------------------------------------------
    logger.info(f"Writing Excel file: {output_path}")
    new_wb = Workbook()
    new_sheet = new_wb.active
    new_sheet.append([
        "Auction ID", "Status", "Minimum Bid", "Open Date",
        "Attorney", "Debt Amount", "Book/Writ", "OPA", "Address",
        "Neighborhood", "Lat", "Lng", "Phila Link", "Bid4Assets Link", "Google Street View"
    ])

    for r in results:
        new_sheet.append([
            r["auction_id"],
            r["status"],
            r["min_bid"],
            r["open_date"],
            r["attorney"],
            r["debt_amount"],
            r["book_writ"],
            r["opa"],
            r["address"],
            r["neighborhood"],
            r["lat"],
            r["lng"],
            r["phila_link"],
            r["bid4assets_link"],
            r["streetview"]
        ])

    new_wb.save(output_path)
    logger.info(f"Excel file saved successfully")

    # -------------------------------------------
    # WRITE GEOJSON
    # -------------------------------------------
    logger.info(f"Writing GeoJSON file: {geojson_path}")
    with open(geojson_path, "w") as f:
        features = []
        for r in results:
            if not r["lat"]:
                continue
            feat = {
                "type": "Feature",
                "geometry": {"type": "Point", "coordinates": [r["lng"], r["lat"]]},
                "properties": r,
            }
            features.append(feat)

        f.write(json.dumps({"type": "FeatureCollection", "features": features}, indent=2))
    logger.info(f"GeoJSON written with {len(features)} features")

    # -------------------------------------------
    # CREATE INTERACTIVE MAP
    # -------------------------------------------
    create_interactive_map(results, map_path)

    logger.info(f"✔ Processing complete!")
    logger.info(f"  - Excel output: {output_path}")
    logger.info(f"  - GeoJSON output: {geojson_path}")
    logger.info(f"  - Map output: {map_path}")


# -------------------------------------------
# ENTRYPOINT
# -------------------------------------------
if __name__ == "__main__":
    import sys
    import os

    if len(sys.argv) < 2:
        logger.error("Usage: python process_auctions.py /path/to/AuctionList.xlsx")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = "AuctionList_Processed.xlsx"
    geojson_file = "AuctionMap.geojson"

    # Map file matches input filename with .html extension
    base_name = os.path.splitext(input_file)[0]
    map_file = f"{base_name}.html"

    logger.info("=" * 60)
    logger.info("AUCTION PROCESSOR STARTED")
    logger.info("=" * 60)

    try:
        asyncio.run(process_file(input_file, output_file, geojson_file, map_file))
    except Exception as e:
        logger.error(f"Fatal error: {e}", exc_info=True)
        sys.exit(1)